import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

# Цвета фона
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def flatten_ordered_dict(d, parent_key='', sep='.', ordered_list=None):
    """Рекурсивно преобразует вложенный словарь в плоский с сохранением порядка"""
    if ordered_list is None:
        ordered_list = []
    
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            flatten_ordered_dict(v, new_key, sep, ordered_list)
        else:
            ordered_list.append((new_key, str(v)))
    return ordered_list

def main():
    # Загрузка и обработка русского JSON с сохранением порядка
    with open('9-ru-RU.json', 'r', encoding='utf-8') as f:
        ru_data = json.load(f)
    flat_ru_ordered = flatten_ordered_dict(ru_data)
    
    # Загрузка английского Excel
    eng_df = pd.read_excel('09-eng.xlsx', header=None)
    eng_dict = dict(zip(eng_df.iloc[:, 0].astype(str), eng_df.iloc[:, 1].astype(str)))
    used_eng_keys = set()

    # Создаем книгу Excel
    wb = Workbook()
    ws_ru = wb.active
    ws_ru.title = "Русские ключи"
    ws_unused_eng = wb.create_sheet("Неиспользованные английские")

    # Заголовки
    ws_ru.append(["Ключ", "Значение RU", "Значение ENG"])
    ws_unused_eng.append(["Ключ ENG", "Значение ENG"])

    # Обрабатываем русские ключи в оригинальном порядке
    for ru_key, ru_value in flat_ru_ordered:
        if ru_key in eng_dict:
            # Совпадение - добавляем английский перевод
            ws_ru.append([ru_key, ru_value, eng_dict[ru_key]])
            # Закрашиваем всю строку зеленым
            for row in ws_ru.iter_rows(min_row=ws_ru.max_row, max_row=ws_ru.max_row):
                for cell in row:
                    cell.fill = GREEN_FILL
            used_eng_keys.add(ru_key)
        else:
            # Нет совпадения - добавляем только русские данные
            ws_ru.append([ru_key, ru_value, ""])
            # Закрашиваем только первые две ячейки красным
            for row in ws_ru.iter_rows(min_row=ws_ru.max_row, max_row=ws_ru.max_row):
                for cell in row[:2]:  # Только ключ и RU значение
                    cell.fill = RED_FILL

    # Добавляем неиспользованные английские ключи
    for eng_key, eng_value in eng_dict.items():
        if eng_key not in used_eng_keys:
            ws_unused_eng.append([eng_key, eng_value])

    # Автонастройка ширины столбцов
    for sheet in [ws_ru, ws_unused_eng]:
        for column in sheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column
            )
            sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

    # Сохраняем файл
    wb.save("9-colored-output.xlsx")
    print(f"Файл создан. Русских ключей: {len(flat_ru_ordered)}, "
          f"использовано английских: {len(used_eng_keys)}, "
          f"неиспользованных английских: {len(eng_dict) - len(used_eng_keys)}")

if __name__ == "__main__":
    main()