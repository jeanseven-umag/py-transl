import json
import openpyxl
from openpyxl.utils import get_column_letter
from copy import deepcopy

def load_json(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_json(data, file_path):
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def set_nested_value(data, key, value):
    keys = key.split('.')
    current = data
    for k in keys[:-1]:
        if k not in current:
            current[k] = {}
        current = current[k]
    current[keys[-1]] = value

def get_nested_value_with_path(data, key):
    keys = key.split('.')
    path = []
    current = data
    
    for k in keys:
        if k in current:
            path.append(k)
            current = current[k]
        else:
            return None, '.'.join(path) if path else None
    
    return current, None

def process_files(xlsx_path, json_path, output_xlsx_path, output_json_path):
    # Загружаем JSON
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    # Создаем копию для английской версии
    en_json_data = deepcopy(json_data)
    
    # Загружаем XLSX
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # Создаем новую книгу для вывода XLSX
    output_wb = openpyxl.Workbook()
    output_ws_matched = output_wb.active
    output_ws_matched.title = "Совпадения"
    output_ws_unmatched = output_wb.create_sheet("Несовпадения")
    
    # Заголовки для совпадений
    output_ws_matched.append(["Ключ", "Значение из JSON (рус)", "Значение из XLSX (англ)"])
    
    # Заголовки для несовпадений
    output_ws_unmatched.append(["Ключ", "Значение из XLSX (англ)", "Путь в JSON где искали"])
    
    # Обрабатываем строки исходного файла
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:  # Пропускаем пустые строки
            continue
            
        key = row[0]
        eng_value = row[1] if len(row) > 1 else None
        
        # Получаем значение из JSON и путь где искали
        ru_value, search_path = get_nested_value_with_path(json_data, key)
        
        if ru_value is not None:
            # Записываем в лист совпадений
            output_ws_matched.append([key, ru_value, eng_value])
            # Обновляем английский JSON
            if eng_value is not None:
                set_nested_value(en_json_data, key, eng_value)
        else:
            # Записываем в лист несовпадений
            output_ws_unmatched.append([key, eng_value, search_path or "Не найден в JSON"])
    
    # Автонастройка ширины столбцов
    for sheet in [output_ws_matched, output_ws_unmatched]:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем результаты
    output_wb.save(output_xlsx_path)
    save_json(en_json_data, output_json_path)
    
    print(f"Обработка завершена. Результаты сохранены в:")
    print(f"- XLSX файл: {output_xlsx_path}")
    print(f"- JSON файл (en-US): {output_json_path}")

if __name__ == "__main__":
    xlsx_input = "09.xlsx"
    json_input = "9-ru-RU.json"
    xlsx_output = "09-output-output.xlsx"
    json_output = "9-en-US-US.json"
    
    process_files(xlsx_input, json_input, xlsx_output, json_output)