import os
import re
from openpyxl import Workbook

def is_russian_text(text):
    """Check if text contains Russian Cyrillic characters"""
    russian_letters = re.compile(r'[а-яА-ЯЁё]')
    return bool(russian_letters.search(text))

def extract_russian_text_from_file(file_path):
    """Extract complete Russian text segments from a PHP file with line numbers"""
    results = []
    with open(file_path, 'r', encoding='utf-8') as file:
        for line_num, line in enumerate(file, start=1):
            # First remove all PHP code to avoid duplicate captures
            line_no_php = re.sub(r'<\?php.*?\?>|<\?.*?\?>', '', line)
            
            # Extract text from HTML tags
            tag_texts = re.findall(r'>([^<]+)<', line_no_php)
            for text in tag_texts:
                text = text.strip()
                if text and is_russian_text(text):
                    results.append((line_num, text))
            
            # Check remaining text that wasn't inside HTML tags
            remaining_text = re.sub(r'<[^>]+>', '', line_no_php).strip()
            if remaining_text and is_russian_text(remaining_text):
                # Split by common delimiters but keep larger segments
                segments = re.split(r'[;{}()\[\],]', remaining_text)
                for segment in segments:
                    segment = segment.strip()
                    if segment and is_russian_text(segment) and segment not in [t for (_, t) in results]:
                        results.append((line_num, segment))
    
    return results

def process_php_files(directory, output_file):
    """Process all PHP files in directory and save results to Excel"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    for filename in os.listdir(directory):
        if filename.endswith('.php'):
            file_path = os.path.join(directory, filename)
            russian_text = extract_russian_text_from_file(file_path)
            
            if russian_text:
                sheet_name = filename.replace('.php', '')[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=sheet_name)
                
                # Write headers
                ws['A1'] = 'Line Number'
                ws['B1'] = 'Russian Text'
                
                # Write data
                for row_num, (line_num, text) in enumerate(russian_text, start=2):
                    ws[f'A{row_num}'] = line_num
                    ws[f'B{row_num}'] = text
                
                # Auto-size columns
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 50
    
    wb.save(output_file)
    print(f"Results saved to {output_file}")

if __name__ == "__main__":
    php_files_directory = 'php_files'
    output_excel_file = 'russian_text_extraction.xlsx'
    
    if not os.path.exists(php_files_directory):
        os.makedirs(php_files_directory)
        print(f"Created directory '{php_files_directory}'. Please place your PHP files there.")
    else:
        process_php_files(php_files_directory, output_excel_file)