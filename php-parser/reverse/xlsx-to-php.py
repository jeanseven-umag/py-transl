import openpyxl
import os
from pathlib import Path

def translate_php_files(xlsx_path, source_dir='ru', target_dir='tj'):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        php_file_name = f"{sheet_name.lower()}.php"
        source_path = os.path.join(source_dir, php_file_name)
        target_path = os.path.join(target_dir, php_file_name)
        
        # Check if source PHP file exists
        if not os.path.exists(source_path):
            print(f"Warning: Source file {source_path} not found for sheet {sheet_name}")
            continue
            
        # Read the source PHP file
        with open(source_path, 'r', encoding='utf-8') as f:
            php_content = f.read()
        
        # Parse the PHP array (this is a simplified approach)
        # In a real implementation, you might want to use a proper PHP parser
        translations = {}
        start = php_content.find('return [') + len('return [')
        end = php_content.rfind('];')
        entries = php_content[start:end].strip().split('\n')
        
        for entry in entries:
            entry = entry.strip().strip(',')
            if not entry:
                continue
            # Split key and value
            parts = entry.split('=>', 1)
            if len(parts) == 2:
                key = parts[0].strip().strip("'\"")
                value = parts[1].strip().strip("'\"")
                translations[key] = value
        
        # Find translations in Excel sheet (column 1 = key, column 5 = translation)
        for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
            if len(row) >= 5 and row[0] in translations:
                translations[row[0]] = row[4] if row[4] else translations[row[0]]  # keep original if no translation
        
        # Generate new PHP content
        new_php_content = "<?php\nreturn [\n"
        for key, value in translations.items():
            # Escape single quotes and preserve newlines
            value = value.replace("'", "\\'").replace('\\"', '"')
            new_php_content += f"    '{key}' => '{value}',\n"
        new_php_content += "];\n"
        
        # Ensure target directory exists
        Path(target_dir).mkdir(parents=True, exist_ok=True)
        
        # Write the translated PHP file
        with open(target_path, 'w', encoding='utf-8') as f:
            f.write(new_php_content)
        
        print(f"Created {target_path}")

if __name__ == "__main__":
    # Example usage
    translate_php_files('tj.xlsx')