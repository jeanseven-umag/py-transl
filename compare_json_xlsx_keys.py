import json
import openpyxl
from openpyxl import Workbook

def load_json_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    keys = []
    values = []
    
    def extract_data(obj, current_key=""):
        if isinstance(obj, dict):
            for key, value in obj.items():
                new_key = f"{current_key}.{key}" if current_key else key
                extract_data(value, new_key)
        else:
            keys.append(current_key)
            values.append(obj)
    
    extract_data(data)
    return keys, values

def load_excel_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    excel_keys = []
    excel_values = []
    
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is not None:
            excel_keys.append(row[0])
            excel_values.append(row[1] if len(row) > 1 else None)
    
    return excel_keys, excel_values

def compare_data(json_keys, json_values, excel_keys, excel_values):
    results = []
    
    # Создаем словари для быстрого поиска
    json_dict = {key: (idx+1, val) for idx, (key, val) in enumerate(zip(json_keys, json_values))}
    excel_dict = {key: (idx+1, val) for idx, (key, val) in enumerate(zip(excel_keys, excel_values))}
    
    # Проверяем все ключи из JSON
    for json_key, json_val in zip(json_keys, json_values):
        json_idx = json_dict[json_key][0]
        
        if json_key in excel_dict:
            excel_idx, excel_val = excel_dict[json_key]
            if abs(json_idx - excel_idx) > 2:
                results.append((
                    json_key, json_key, 
                    json_idx, excel_idx, 
                    excel_val, json_val
                ))
        else:
            # Ищем похожие ключи в Excel
            similar_keys = [k for k in excel_keys 
                           if k.startswith(json_key.split('.')[0]) or json_key.split('.')[0] in k]
            
            if similar_keys:
                for similar_key in similar_keys[:1]:  # Берем первый похожий ключ
                    excel_idx, excel_val = excel_dict[similar_key]
                    results.append((
                        json_key, similar_key,
                        json_idx, excel_idx,
                        excel_val, json_val
                    ))
            else:
                results.append((
                    json_key, "",
                    json_idx, "",
                    "", json_val
                ))
    
    # Проверяем ключи из Excel, которых нет в JSON
    for excel_key in excel_keys:
        if excel_key not in json_dict:
            excel_idx, excel_val = excel_dict[excel_key]
            results.append((
                "", excel_key,
                "", excel_idx,
                excel_val, ""
            ))
    
    return results

def save_results_to_excel(results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mismatches"
    
    # Заголовки столбцов
    headers = [
        "JSON Key", "Excel Key", 
        "JSON Index", "Excel Index", 
        "Excel Value", "JSON Value"
    ]
    ws.append(headers)
    
    # Заполняем данные
    for row in results:
        ws.append(row)
    
    # Настраиваем ширину столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_file)

def main():
    json_file = "9-ru-RU.json"
    excel_file = "09.xlsx"
    output_file = "09_output.xlsx"
    
    # Загружаем данные
    json_keys, json_values = load_json_data(json_file)
    excel_keys, excel_values = load_excel_data(excel_file)
    
    # Сравниваем данные
    results = compare_data(json_keys, json_values, excel_keys, excel_values)
    
    # Сохраняем результаты
    save_results_to_excel(results, output_file)
    
    print(f"Сравнение завершено. Результаты сохранены в {output_file}")
    print(f"Всего найдено несоответствий: {len(results)}")

if __name__ == "__main__":
    main()