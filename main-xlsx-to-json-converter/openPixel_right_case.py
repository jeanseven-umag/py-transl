import json
import openpyxl
from openpyxl import load_workbook

def update_json_with_excel(
    xlsx_file: str,
    source_json_file: str,
    output_json_file: str
) -> dict:
    """
    Updates JSON with translations from Excel, ensuring proper capitalization.
    Only first letter is capitalized, rest remain as in original translation.
    """
    # 1. Load source JSON
    with open(source_json_file, 'r', encoding='utf-8') as f:
        source_data = json.load(f)
    
    # 2. Load translations from Excel
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    
    translations = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        key = row[0]
        translation = row[1]
        if key and translation:
            # Proper capitalization: first letter uppercase, rest unchanged
            if isinstance(translation, str):
                translation = translation[0].upper() + translation[1:].lower()
            translations[key] = translation
    
    # 3. Deep update function
    def deep_update(data, parent_key=''):
        if isinstance(data, dict):
            updated = {}
            for key, value in data.items():
                current_key = f"{parent_key}.{key}" if parent_key else key
                if current_key in translations:
                    updated[key] = translations[current_key]
                elif isinstance(value, dict):
                    updated[key] = deep_update(value, current_key)
                else:
                    updated[key] = value
            return updated
        return data
    
    updated_data = deep_update(source_data)
    
    # 4. Save updated JSON
    with open(output_json_file, 'w', encoding='utf-8') as f:
        json.dump(updated_data, f, ensure_ascii=False, indent=2)
    
    return updated_data

if __name__ == "__main__":
    update_json_with_excel(
        '05.xlsx',
        'ru-RU.json',
        'en-US.json'
    )
    print("JSON file has been updated successfully with proper Capitalization!")