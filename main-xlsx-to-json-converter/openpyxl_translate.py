import json
import openpyxl
from openpyxl import load_workbook

def update_json_with_excel(
    xlsx_file: str,
    source_json_file: str,
    output_json_file: str
) -> dict:
    # 1. Load the source JSON file
    with open(source_json_file, 'r', encoding='utf-8') as f:
        source_data = json.load(f)
    
    # 2. Load translations from Excel (column A = keys, column B = translations)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    
    translations = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        key = row[0]  # Column A = key
        translation = row[1]  # Column B = translation
        if key and translation:
            translations[key] = translation
    
    # 3. Deep update function for nested JSON
    def deep_update(data, parent_key=''):
        if isinstance(data, dict):
            updated = {}
            for key, value in data.items():
                current_key = f"{parent_key}.{key}" if parent_key else key
                
                # Case 1: Direct key match (e.g., "transfers-history")
                if current_key in translations:
                    updated[key] = translations[current_key]
                # Case 2: Nested dictionary (e.g., "create" -> "create-transfer")
                elif isinstance(value, dict):
                    updated[key] = deep_update(value, current_key)
                # Case 3: Keep original if no translation
                else:
                    updated[key] = value
            return updated
        return data
    
    updated_data = deep_update(source_data)
    
    # 4. Save the updated JSON
    with open(output_json_file, 'w', encoding='utf-8') as f:
        json.dump(updated_data, f, ensure_ascii=False, indent=2)
    
    return updated_data

# Example usage
if __name__ == "__main__":
    updated_data = update_json_with_excel(
        '6_Home_converted.xlsx',
        'ru-RU.json',
        '6_en-US.json'
    )
    print("JeanSON file has been updated successfully!")