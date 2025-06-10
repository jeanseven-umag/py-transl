import json
import openpyxl
from openpyxl import Workbook

def json_to_excel(json_file: str, excel_file: str):
    """
    Конвертирует JSON в Excel (ключи в колонке A, значения в колонке B)
    Обрабатывает все уровни вложенности, сохраняя полный путь к ключу
    """
    
    # 1. Загружаем JSON
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 2. Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"
    
    # 3. Заголовки
    ws['A1'] = "Key"
    ws['B1'] = "Translation"
    
    # 4. Рекурсивная функция для извлечения всех ключей и значений
    def extract_items(data, parent_key='', row_num=2):
        if isinstance(data, dict):
            for key, value in data.items():
                current_key = f"{parent_key}.{key}" if parent_key else key
                
                if isinstance(value, dict):
                    row_num = extract_items(value, current_key, row_num)
                else:
                    ws.cell(row=row_num, column=1, value=current_key)
                    ws.cell(row=row_num, column=2, value=value)
                    row_num += 1
        elif isinstance(data, str):
            ws.cell(row=row_num, column=1, value=parent_key)
            ws.cell(row=row_num, column=2, value=data)
            row_num += 1
        return row_num
    
    # 5. Заполняем данные
    extract_items(data)
    
    # 6. Сохраняем Excel
    wb.save(excel_file)
    print(f"Файл {excel_file} успешно создан! Всего строк: {ws.max_row - 1}")

# Пример использования
if __name__ == "__main__":
    json_to_excel(
        json_file="ru-RU.json",
        excel_file="03.xlsx"
    )